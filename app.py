from flask import Flask, request, send_file
import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

# Folders
DATA_FOLDER = "data"
ARCHIVE_FOLDER = "archive"
COLUMNS = [
    "Tenant Name", "Tenant Code", "Golive AM", "Go Live Mgr",
    "Current_Status", "Dashboard Status", "Remarks"
]

# Generate Excel file name based on current week
def get_excel_filename():
    week = datetime.now().isocalendar().week
    year = datetime.now().year
    return f"dashboard_{year}-W{week}.xlsx"

@app.route("/")
def form():
    with open("dash.htm", "r", encoding="utf-8") as f:
        return f.read()

@app.route("/submit", methods=["POST"])
def submit():
    os.makedirs(DATA_FOLDER, exist_ok=True)
    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

    new_file = os.path.join(DATA_FOLDER, get_excel_filename())
    print("➡️ Saving to:", new_file)

    # Archive old files
    for fname in os.listdir(DATA_FOLDER):
        if fname.endswith(".xlsx") and fname != os.path.basename(new_file):
            shutil.move(os.path.join(DATA_FOLDER, fname), os.path.join(ARCHIVE_FOLDER, fname))
            print("📦 Archived:", fname)

    # Read form input
    new_data = {
        "Tenant Name": request.form.get("tenant_name"),
        "Tenant Code": request.form.get("tenant_code"),
        "Golive AM": request.form.get("golive_am"),
        "Go Live Mgr": request.form.get("golive_mgr"),
        "Current_Status": request.form.get("status"),
        "Dashboard Status": request.form.get("dashboard_status"),
        "Remarks": request.form.get("remarks")
    }

    print("📝 Form data received:", new_data)

    try:
        if os.path.exists(new_file):
            df = pd.read_excel(new_file)
            df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)
        else:
            df = pd.DataFrame([new_data], columns=COLUMNS)

        df = df[COLUMNS]
        df.to_excel(new_file, index=False)
        apply_styling(new_file)
        print("✅ Excel file written and styled.")

    except Exception as e:
        print("❌ Error while writing Excel:", e)
        return f"<h3>Error: {e}</h3>"

    return f"<h3>✅ Dashboard updated! <a href='/'>Go Back</a></h3><br><a href='/download'>📥 Download Excel</a>"

@app.route("/download")
def download():
    excel_file = os.path.join(DATA_FOLDER, get_excel_filename())
    if os.path.exists(excel_file):
        return send_file(excel_file, as_attachment=True)
    else:
        return "<h3>⚠️ No dashboard file available for this week yet.</h3>"

def apply_styling(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Style rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left")

    # Auto column width
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 5

    wb.save(file_path)

# For local or Render deployment
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
